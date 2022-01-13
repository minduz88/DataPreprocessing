import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8')

import re
import requests
import json

print("Start getting to tags..")

from openpyxl import load_workbook
from openpyxl import Workbook

load_wb1 = load_workbook("C:\\Files\\North.xlsx", data_only=True)
write_wb=Workbook()

write_ws=write_wb.create_sheet("sheet1")
write_ws=write_wb.active

# get a sheet name
load_ws1 = load_wb1['Sheet1']
print(load_ws1.max_row)

tagType=('UNIT', 'INSTLTN', 'EQP', 'PRSN', 'LOCTN', 'ACTVT')

write_ws.cell(1,1,'번호')
write_ws.cell(1,2,'id')
write_ws.cell(1,3,'문서내용')
write_ws.cell(1,4,'태그')
write_ws.cell(1,5,'태그내용')
m=2
for i in range(2, load_ws1.max_row+1):
            strCount=load_ws1.cell(i,3).value
            str1 = re.findall('<UNIT>(.*?)</UNIT>|<INSTLTN>(.*?)</INSTLTN>|<EQP>(.*?)</EQP>|<PRSN>(.*?)</PRSN>|<LOCTN>(.*?)</LOCTN>|<ACTVT>(.*?)</ACTVT>', strCount)

            print(str1)
            print(len(str1))
            for j in range(0, len(str1)):
                        print(str1[j])
                        arrStrK = str1[j]
                        for k in range(0, len(arrStrK)):
                                    print(arrStrK[k])
                                    print(len(arrStrK[k]))
                                    if len(arrStrK[k]) > 0:
                                                write_ws.cell(m,1, load_ws1.cell(i,1).value)
                                                write_ws.cell(m,2, load_ws1.cell(i,2).value)
                                                write_ws.cell(m,3, load_ws1.cell(i,3).value)
                                                write_ws.cell(m,4, tagType[k])
                                                write_ws.cell(m,5, arrStrK[k])
                                                m=m+1

write_wb.save('C:\\Files\\North_result.xlsx')

print("End.")


                                                                                      